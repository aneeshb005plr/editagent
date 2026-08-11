"""
app/documents/xlsx_parser.py

Parses .xlsx files into the common ParsedDocument representation
(app/documents/base.py). Built and verified against openpyxl 3.1.5.

SCOPE, as confirmed in discovery: every tab ("multiple pages" = every
sheet) is reviewed; only STRING-TYPED cells are treated as reviewable
text - formulas and numeric values have nothing for a grammar/tone/
risk-language check to act on. openpyxl's cell.data_type == 's'
directly identifies string cells even in read_only mode - confirmed
against a real fixture (data_only=False), NOT filtering by checking
`isinstance(value, str)`, because a formula's raw value under
data_only=False is itself a string like "=A2*2" and would be
wrongly treated as prose without the data_type check.

read_only=True used throughout for CELL text - confirmed memory-
efficient mode per prior-project precedent; NOT yet load-tested at
real large multi-tab workbook scale.

IMAGES/CHARTS: openpyxl's read-only mode does not expose these, and
reloading the whole workbook non-read-only just to find them would
reintroduce the exact memory problem read_only=True exists to avoid
at 100MB scale. Instead, images/charts are found by reading the
.xlsx ZIP CONTAINER directly (an .xlsx is a zip of XML parts) and
walking the real, confirmed relationship chain: xl/workbook.xml
(sheet name -> r:id) -> xl/_rels/workbook.xml.rels (r:id -> sheetN.
xml) -> xl/worksheets/_rels/sheetN.xml.rels (-> drawingK.xml, if
present) -> drawingK.xml (anchor cell position + per-shape rId) ->
drawingK.xml.rels (rId -> image relationship, giving the real media
path, OR chart relationship, no bytes needed). This whole chain was
verified against real fixtures generated with openpyxl (both an
embedded image and an embedded chart), not assumed from
documentation - see conversation history for the raw XML inspected.
This adds one more read of the zip container but never loads cell
data non-read-only, so the 100MB memory profile for cell scanning is
unaffected.

EXPLICITLY OUT OF SCOPE, not silently missing:
- Chartsheets (a whole worksheet that IS a chart, a distinct OOXML
  part type from a normal worksheet) are not covered -
  _sheet_name_to_xml_path only resolves relationships of type
  "worksheet", so a chart living in a chartsheet is never flagged.
- Cell comments/notes (stored in a separate CommentSheet part) are
  not read at all, even though they can contain reviewable prose -
  the read-only cell-scanning pass never touches this part.
- Linked (not embedded) images are detected (via TargetMode=
  "External") and flagged as NOT_APPLICABLE, but their content is
  never fetched - there's no local media to read.
"""

from __future__ import annotations

import io
import logging
import mimetypes
import zipfile

import openpyxl
from lxml import etree
from openpyxl.cell.read_only import EmptyCell
from openpyxl.utils import get_column_letter

from app.documents.base import (
    ContentBlock,
    ContentKind,
    ExtractionStatus,
    Location,
    ParsedDocument,
    UnsupportedItem,
    UnsupportedKind,
)

logger = logging.getLogger("app.documents.xlsx_parser")

_PKG_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
_SPREADSHEET_DRAWING_NS = (
    "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
)
_REL_TYPE_IMAGE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
_REL_TYPE_CHART = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart"
_REL_TYPE_DRAWING = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
_REL_TYPE_WORKSHEET = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"


def _parse_rels(zf: zipfile.ZipFile, rels_path: str) -> dict[str, tuple[str, str]]:
    """Returns {r_id: (rel_type, resolved_target_path)} for a given
    .rels part, or {} if that part doesn't exist (e.g. a sheet with
    no drawing at all - the normal, common case)."""

    if rels_path not in zf.namelist():
        return {}

    tree = etree.fromstring(zf.read(rels_path))
    base_dir = rels_path.rsplit("/_rels/", 1)[0]
    result: dict[str, tuple[str, str]] = {}

    for rel in tree.findall(f"{_PKG_REL_NS}Relationship"):
        r_id = rel.get("Id")
        rel_type = rel.get("Type")
        target = rel.get("Target")
        target_mode = rel.get("TargetMode")  # "External" for linked (non-embedded) media

        if not target or not r_id:
            # Malformed relationship entry - skip rather than crash
            # the whole pass with an AttributeError on target.startswith()
            # a few lines down.
            continue

        if target_mode == "External":
            # A linked (not embedded) image/media reference - the
            # Target is an external URI, not a path inside this zip,
            # so there's nothing for zf.read() to resolve. Recorded
            # with a synthetic marker rather than a real zip path;
            # the caller checks for this explicitly.
            result[r_id] = (rel_type, f"__EXTERNAL__:{target}")
            continue

        if target.startswith("/"):
            resolved = target.lstrip("/")
        else:
            resolved = f"{base_dir}/{target}"
            # Normalize any "../" segments.
            parts: list[str] = []
            for segment in resolved.split("/"):
                if segment == "..":
                    if parts:
                        parts.pop()
                else:
                    parts.append(segment)
            resolved = "/".join(parts)

        result[r_id] = (rel_type, resolved)

    return result


def _cell_ref_from_anchor(anchor_elem) -> str | None:
    from_elem = anchor_elem.find(f"{_SPREADSHEET_DRAWING_NS}from")
    if from_elem is None:
        return None

    col_elem = from_elem.find(f"{_SPREADSHEET_DRAWING_NS}col")
    row_elem = from_elem.find(f"{_SPREADSHEET_DRAWING_NS}row")
    if col_elem is None or row_elem is None:
        return None

    # Anchor col/row are 0-indexed in the XML; cell references and
    # get_column_letter are both 1-indexed - confirmed by generating
    # a fixture anchored at "C3" and checking the raw XML produced
    # <col>2</col><row>2</row>.
    col = int(col_elem.text) + 1
    row = int(row_elem.text) + 1
    return f"{get_column_letter(col)}{row}"


def _find_drawing_refs(drawing_xml: bytes) -> list[tuple[str, str | None]]:
    """Returns [(r_id, cell_reference)] for every picture or chart
    graphicFrame anchored in this drawing part. r_id still needs
    resolving against the drawing's OWN .rels part to know whether
    it's an image or a chart."""

    tree = etree.fromstring(drawing_xml)
    refs: list[tuple[str, str | None]] = []

    for anchor in list(tree):
        # oneCellAnchor or twoCellAnchor - only "from" is needed.
        cell_ref = _cell_ref_from_anchor(anchor)

        # Picture: .//pic/blipFill/blip/@r:embed
        for blip in anchor.iter(
            "{http://schemas.openxmlformats.org/drawingml/2006/main}blip"
        ):
            r_id = blip.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
            )
            if r_id:
                refs.append((r_id, cell_ref))

        # Chart: .//graphicFrame//c:chart/@r:id
        for chart_ref in anchor.iter(
            "{http://schemas.openxmlformats.org/drawingml/2006/chart}chart"
        ):
            r_id = chart_ref.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            )
            if r_id:
                refs.append((r_id, cell_ref))

    return refs


def _extract_drawings_for_sheet(
    zf: zipfile.ZipFile,
    sheet_xml_path: str,
    sheet_name: str,
) -> list[UnsupportedItem]:
    sheet_dir, sheet_file = sheet_xml_path.rsplit("/", 1)
    sheet_rels_path = f"{sheet_dir}/_rels/{sheet_file}.rels"
    sheet_rels = _parse_rels(zf, sheet_rels_path)

    drawing_targets = [
        target for rel_type, target in sheet_rels.values()
        if rel_type == _REL_TYPE_DRAWING
    ]

    items: list[UnsupportedItem] = []

    for drawing_path in drawing_targets:
        if drawing_path not in zf.namelist():
            continue

        drawing_dir, drawing_file = drawing_path.rsplit("/", 1)
        drawing_rels_path = f"{drawing_dir}/_rels/{drawing_file}.rels"
        drawing_rels = _parse_rels(zf, drawing_rels_path)

        for r_id, cell_ref in _find_drawing_refs(zf.read(drawing_path)):
            if r_id not in drawing_rels:
                continue

            rel_type, target_path = drawing_rels[r_id]
            location = Location(sheet_name=sheet_name, cell_reference=cell_ref)

            if rel_type == _REL_TYPE_IMAGE:
                if target_path.startswith("__EXTERNAL__:"):
                    # Linked (not embedded) image - real, if uncommon,
                    # case: no local bytes exist to extract, confirmed
                    # via the OOXML TargetMode="External" attribute.
                    items.append(
                        UnsupportedItem(
                            kind=UnsupportedKind.IMAGE,
                            location=location,
                            note="linked (not embedded) image - no local bytes to extract",
                            extraction_status=ExtractionStatus.NOT_APPLICABLE,
                        )
                    )
                    continue

                try:
                    raw_bytes = zf.read(target_path)
                    guessed_type, _ = mimetypes.guess_type(target_path)
                    content_type = guessed_type or f"image/{target_path.rsplit('.', 1)[-1].lower()}"
                    items.append(
                        UnsupportedItem(
                            kind=UnsupportedKind.IMAGE,
                            location=location,
                            note=f"content_type={content_type}",
                            raw_bytes=raw_bytes,
                            extraction_status=ExtractionStatus.PENDING,
                        )
                    )
                except KeyError:
                    logger.warning(
                        "Could not read image media at %s for sheet %s",
                        target_path,
                        sheet_name,
                        exc_info=True,
                    )
                    items.append(
                        UnsupportedItem(
                            kind=UnsupportedKind.IMAGE,
                            location=location,
                            note="image bytes unavailable",
                            extraction_status=ExtractionStatus.FAILED,
                        )
                    )

            elif rel_type == _REL_TYPE_CHART:
                items.append(
                    UnsupportedItem(
                        kind=UnsupportedKind.CHART,
                        location=location,
                        note="chart data/visual layout not reviewed in current scope",
                        extraction_status=ExtractionStatus.NOT_APPLICABLE,
                    )
                )

    return items


def _sheet_name_to_xml_path(zf: zipfile.ZipFile) -> dict[str, str]:
    """Resolves the full chain: sheet name -> r:id (workbook.xml) ->
    worksheet target (workbook.xml.rels) -> zip-internal sheetN.xml
    path. Needed because sheet DECLARATION ORDER in workbook.xml does
    not reliably match sheetN.xml FILE NUMBERING - r:id is the only
    correct way to join them, confirmed against real fixture XML."""

    workbook_xml = etree.fromstring(zf.read("xl/workbook.xml"))
    main_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    r_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

    name_to_rid = {
        sheet.get("name"): sheet.get(f"{r_ns}id")
        for sheet in workbook_xml.find(f"{main_ns}sheets")
    }

    workbook_rels = _parse_rels(zf, "xl/_rels/workbook.xml.rels")

    result: dict[str, str] = {}
    for name, r_id in name_to_rid.items():
        if r_id in workbook_rels:
            rel_type, target = workbook_rels[r_id]
            if rel_type == _REL_TYPE_WORKSHEET:
                result[name] = target

    return result


def parse_xlsx(file_bytes: bytes, filename: str) -> ParsedDocument:
    """Parses an .xlsx file's readable content into a ParsedDocument.

    File-size validation is the DISPATCHER's responsibility (single
    source of truth against settings.MAX_FILE_SIZE_MB) - this
    function assumes it's already been called with an acceptable
    size and focuses purely on content extraction.
    """

    workbook = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        read_only=True,
        data_only=False,
        # data_only=False deliberately - see module docstring on why
        # we need the data_type check rather than raw values.
    )

    parsed = ParsedDocument(source_filename=filename, file_type="xlsx")

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell, EmptyCell):
                    continue

                if cell.data_type not in ("s", "inlineStr"):
                    # Formula ('f') or numeric ('n') - not reviewable
                    # prose, skip per confirmed scope. Widened to also
                    # accept 'inlineStr' defensively: confirmed by
                    # reading openpyxl 3.1.5's actual parse_cell()
                    # source that inlineStr is already remapped to
                    # 's' in THIS version's read-only mode, so this
                    # branch is currently unreachable for inlineStr -
                    # kept as hardening against a future openpyxl
                    # version regressing that behavior, not a fix for
                    # an active bug in 3.1.5.
                    continue

                raw = cell.value
                # str(raw) rather than (cell.value or "") - confirmed
                # CellRichText (openpyxl.cell.rich_text.CellRichText)
                # is a list subclass with NO .strip() method; if a
                # rich-text cell value were ever encountered despite
                # the data_type=='s' filter, (cell.value or "").strip()
                # would raise AttributeError. str(raw) degrades
                # gracefully instead of crashing the whole parse.
                text = str(raw).strip() if raw is not None else ""
                if not text:
                    continue

                col_letter = get_column_letter(cell.column)
                cell_ref = f"{col_letter}{cell.row}"

                parsed.blocks.append(
                    ContentBlock(
                        text=text,
                        kind=ContentKind.SPREADSHEET_CELL,
                        location=Location(
                            sheet_name=worksheet.title,
                            cell_reference=cell_ref,
                        ),
                    )
                )

    sheet_count = len(workbook.sheetnames)
    workbook.close()

    # Second pass, zip-container-level, for images/charts - see
    # module docstring for why this is separate from the read_only
    # cell pass above rather than a single combined load.
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            sheet_paths = _sheet_name_to_xml_path(zf)
            for sheet_name, sheet_xml_path in sheet_paths.items():
                parsed.unsupported_items.extend(
                    _extract_drawings_for_sheet(zf, sheet_xml_path, sheet_name)
                )
    except (zipfile.BadZipFile, KeyError, etree.XMLSyntaxError):
        logger.warning(
            "Could not read drawing/chart relationships for %s - "
            "text content was still parsed normally, images/charts "
            "in this file just won't be flagged",
            filename,
            exc_info=True,
        )

    logger.info(
        "Parsed %s: %d text blocks, %d unsupported items across %d sheets (%.0f chars)",
        filename,
        len(parsed.blocks),
        len(parsed.unsupported_items),
        sheet_count,
        parsed.total_char_count,
    )

    return parsed